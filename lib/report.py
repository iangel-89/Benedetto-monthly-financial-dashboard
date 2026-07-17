# ==============================================================================
# EXECUTIVE SUMMARY TABLE + EXCEL EXPORT
# Structured, numbers-first summary used both for the on-screen "key figures"
# table and as the factual grounding for the AI narrative — the AI is never
# asked to compute anything, only to explain numbers this module already
# computed deterministically.
# ==============================================================================
from __future__ import annotations

import io
import datetime as dt

import pandas as pd

from .engine import CFG


def money(v) -> str:
    return "n/a" if pd.isna(v) else f"${v:,.0f}"


def pct(v, dp: int = 1) -> str:
    return "n/a" if pd.isna(v) else f"{v:,.{dp}f}%"


def executive_summary(M: pd.DataFrame, entity: str) -> pd.DataFrame:
    """The one table a CFO (or a Gemini prompt) actually reads: totals, closing
    position, and a plain verdict on each figure."""
    first, last = M.index[0], M.index[-1]
    n = len(M.index)

    flow = lambda c: M[c].sum()
    stock = lambda c: M[c].iloc[-1]

    sales_t, ebitda_t, ebit_t, ni_t = flow("Sales"), flow("EBITDA"), flow("EBIT"), flow("Net Income")
    ocf_t, capex_t = flow("Operating Cash Flow"), flow("Capex")
    nd, eq, ce = stock("Net Debt"), stock("Equity"), stock("Capital Employed")

    roce = M["ROCE %"].iloc[-1]
    roe = M["ROE %"].iloc[-1]
    nd_ebitda = M["Net Debt / EBITDA (x)"].iloc[-1]
    i_rate = M["After-tax Cost of Debt %"].iloc[-1]

    def verdict(ok, good, bad):
        return good if ok else bad

    rows = [
        ("Company", entity, ""),
        ("Period analysed", f"{first} to {last}  ({n} months)", ""),
        ("", "", ""),
        ("-- WEALTH CREATION --", "", ""),
        ("Total revenue", money(sales_t), ""),
        ("Revenue growth, first to last month",
         pct((M['Sales'][last] / M['Sales'][first] - 1) * 100 if abs(M['Sales'][first]) > 1e-9 else float('nan')),
         verdict(M["Sales"][last] >= M["Sales"][first], "Growing", "Shrinking")),
        ("EBITDA (cash operating profit)", money(ebitda_t), f"Margin {pct(ebitda_t / sales_t * 100 if sales_t else float('nan'))}"),
        ("Net income", money(ni_t), f"Margin {pct(ni_t / sales_t * 100 if sales_t else float('nan'))}"),
        ("Costs vs revenue trend",
         f"{M['Scissors Gap (pts)'][last]:+,.0f} pts" if pd.notna(M['Scissors Gap (pts)'][last]) else "n/a",
         verdict(M["Scissors Gap (pts)"][last] >= 0 if pd.notna(M["Scissors Gap (pts)"][last]) else False,
                 "Revenue outgrowing costs", "Costs outgrowing revenue")),
        ("", "", ""),
        ("-- INVESTMENT --", "", ""),
        ("Capital employed (closing)", money(ce), "Net fixed assets + operating working capital"),
        ("Capital spending (capex)", money(capex_t), f"{pct(capex_t / sales_t * 100 if sales_t else float('nan'))} of revenue"),
        ("", "", ""),
        ("-- FINANCING --", "", ""),
        ("Cash from operations", money(ocf_t), verdict(ocf_t > 0, "Self-funding", "Burning cash")),
        ("Free cash flow (after capex)", money(ocf_t - capex_t),
         verdict(ocf_t - capex_t > 0, "Self-financed growth", "Needs external funding")),
        ("Net debt (closing)", money(nd), ""),
        ("Net debt / EBITDA", f"{nd_ebitda:,.2f}x" if pd.notna(nd_ebitda) else "n/a",
         verdict(pd.notna(nd_ebitda) and nd_ebitda < CFG.ND_EBITDA_CEILING,
                 f"Below the {CFG.ND_EBITDA_CEILING:.0f}x caution line",
                 f"Above the {CFG.ND_EBITDA_CEILING:.0f}x caution line")),
        ("Equity (closing)", money(eq), verdict(eq > 0, "Positive", "Negative — solvency risk")),
        ("", "", ""),
        ("-- RETURNS --", "", ""),
        ("Return on capital employed (ROCE)", pct(roce), ""),
        ("Return on equity (ROE)", pct(roe) if pd.notna(roe) else "n/a — equity too thin to measure", ""),
        ("Is debt helping or hurting returns?",
         verdict(pd.notna(roce) and pd.notna(i_rate) and roce > i_rate,
                 "Helping — return beats the cost of debt",
                 "Hurting — cost of debt exceeds the return"), ""),
    ]
    return pd.DataFrame(rows, columns=["Metric", "Value", "Note"])


STAGE_MAP = {
    "Wealth Creation": ["Sales", "COGS", "Gross Profit", "Operating Expenses (core)",
                         "D&A", "EBITDA", "EBIT", "Interest Expense", "Net Income",
                         "Gross Margin %", "EBITDA Margin %", "EBIT Margin %", "Net Margin %",
                         "Sales Growth %", "COGS Growth %", "Opex Growth %"],
    "Investment Policy": ["Fixed Assets", "Operating Working Capital", "Working Capital (CA-CL)",
                           "Capital Employed", "Capex", "Capex / Sales %", "Capex / D&A (x)",
                           "WC Turnover (x)", "WC in Days of Sales", "DSO (days)", "DIO (days)",
                           "DPO (days)", "Cash Conversion Cycle (days)", "Current Ratio", "Quick Ratio"],
    "Financing Policy": ["Operating Cash Flow", "Investing Cash Flow", "Financing Cash Flow",
                          "Free Cash Flow", "Net Change in Cash", "Cash", "Gross Debt", "Net Debt",
                          "Equity", "Net Debt / EBITDA (x)", "Gearing (ND/Equity) %",
                          "Interest Cover (x)", "OCF / Net Debt %"],
    "Returns & Leverage": ["NOPAT", "ROCE %", "ROE %", "After-tax Cost of Debt %",
                            "Leverage Multiplier (ND/E)", "Leverage Effect (pts)"],
}


# Chart images embedded per sheet — each is a PNG render of the exact same
# Plotly figure shown on the dashboard (via lib.charts + kaleido), so the
# take-home report matches what was actually reviewed on screen instead of a
# separately-styled native Excel chart. Detailed mode places the full
# Vernimmen-stage breakdown on each stage sheet; Simple mode places the four
# plain-language charts on the Executive Summary sheet only, leaving the
# stage sheets (which use Vernimmen terminology) as data tables only.
DETAIL_CHART_FUNCS = {
    "Wealth Creation": ["detail_revenue_momentum", "detail_scissors", "detail_common_size",
                         "detail_variance", "detail_margin_walk"],
    "Investment Policy": ["detail_capital_employed", "detail_working_capital",
                           "detail_cash_conversion_cycle", "detail_capex_da"],
    "Financing Policy": ["detail_net_debt", "detail_leverage", "detail_operating_cf",
                          "detail_cf_allocation"],
    "Returns & Leverage": ["detail_returns"],
}
SIMPLE_CHART_FUNCS = ["simple_health_meters", "simple_revenue_profit",
                       "simple_cash_trend", "simple_revenue_bridge"]


def _fmt_generated() -> str:
    return dt.datetime.now().strftime("%B %d, %Y")


def _write_cover(ws, entity: str, M: pd.DataFrame, has_ai: bool, view: str = "detailed"):
    from openpyxl.styles import Font, Alignment

    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 70

    ws["B2"] = "Financial Health Report"
    ws["B2"].font = Font(bold=True, size=26, color="0B2540")
    ws["B3"] = "Vernimmen four-stage analysis"
    ws["B3"].font = Font(size=12, color="64748B")

    ws["B6"] = entity
    ws["B6"].font = Font(bold=True, size=16, color="0B2540")
    ws["B7"] = f"Period analysed:  {M.index[0]} to {M.index[-1]}   ({len(M.index)} months)"
    ws["B7"].font = Font(size=11, color="52514E")
    ws["B8"] = f"Report generated:  {_fmt_generated()}"
    ws["B8"].font = Font(size=11, color="52514E")
    version_label = ("Simple — plain-language summary" if view == "simple"
                      else "Detailed — full Vernimmen breakdown")
    ws["B9"] = f"Report version:  {version_label}"
    ws["B9"].font = Font(size=11, color="52514E")

    ws["B11"] = "Contents"
    ws["B11"].font = Font(bold=True, size=13, color="0B2540")

    contents = []
    if has_ai:
        contents.append("AI Summary — plain-English executive narrative")
    contents.append("Executive Summary — key figures and read-across")
    contents += [
        "Wealth Creation — revenue, margins, cost dynamics",
        "Investment Policy — capital employed, working capital, liquidity",
        "Financing Policy — debt, cash flow, leverage",
        "Returns & Leverage — ROCE, ROE, leverage effect",
        "Data Quality Log — assumptions and flags to review",
        "Line Item Mapping — which line in your file fed each figure",
    ]
    r = 12
    for item in contents:
        ws.cell(row=r, column=2, value=f"•  {item}").font = Font(size=11, color="0B2540")
        r += 1

    note_r = r + 2
    ws.cell(row=note_r, column=2,
            value=("Figures are computed deterministically from your uploaded statements. "
                   "The AI narrative only explains those figures — it never invents numbers."
                   if has_ai else
                   "No AI narrative was included in this export. Provide a Gemini API key in the "
                   "app to add a plain-English executive summary.")
            ).font = Font(size=9, italic=True, color="8A8781")


def _write_ai_sheet(ws, ai_summary: str, ai_model: str, entity: str):
    from openpyxl.styles import Font, Alignment

    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 100

    ws["B2"] = "AI Executive Summary"
    ws["B2"].font = Font(bold=True, size=18, color="0B2540")
    subtitle = entity + (f"   ·   generated by {ai_model}" if ai_model else "")
    ws["B3"] = subtitle
    ws["B3"].font = Font(size=10, color="64748B")

    r = 5
    for raw in ai_summary.split("\n"):
        line = raw.strip()
        if not line:
            r += 1
            continue
        cell = ws.cell(row=r, column=2)
        if line.startswith("####") or line.startswith("###"):
            cell.value = line.lstrip("#").strip()
            cell.font = Font(bold=True, size=12, color="0B2540")
        else:
            cell.value = line.replace("**", "").replace("__", "").replace("*", "")
            cell.font = Font(size=11, color="1a1a1a")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = max(16, 15 * (len(cell.value) // 95 + 1))
        r += 1


def _fig_to_png(fig, width: int = 860, scale: int = 2) -> bytes:
    """Render a Plotly figure to PNG bytes at its own configured aspect ratio
    (brand_layout sets height; width defaults to Plotly's standard 700px
    unless overridden here), scaled 2x for a crisp look at Excel's display size."""
    height = fig.layout.height or 420
    return fig.to_image(format="png", width=width, height=height, scale=scale)


def _embed_chart_images(ws, anchor_row: int, figs: list, full_width_first: bool = False):
    """Place PNG renders of the given Plotly figures below anchor_row: two per
    row in a grid, or — when full_width_first is set — the first figure alone
    spanning a full row (used for the wide health-meters gauge trio)."""
    from openpyxl.drawing.image import Image as XLImage

    RENDER_W = 860
    HALF_DISP_W, HALF_STEP = 430, 17
    FULL_DISP_W, FULL_STEP = 890, 15

    row = anchor_row
    start = 0
    if full_width_first and figs:
        fig = figs[0]
        png = _fig_to_png(fig, width=RENDER_W * 2)
        img = XLImage(io.BytesIO(png))
        img.width = FULL_DISP_W
        img.height = int((fig.layout.height or 420) * FULL_DISP_W / (RENDER_W * 2))
        ws.add_image(img, f"B{row}")
        row += FULL_STEP
        start = 1

    col_letters = ["B", "L"]
    for i, fig in enumerate(figs[start:]):
        png = _fig_to_png(fig, width=RENDER_W)
        img = XLImage(io.BytesIO(png))
        img.width = HALF_DISP_W
        img.height = int((fig.layout.height or 420) * HALF_DISP_W / RENDER_W)
        col = col_letters[i % 2]
        r = row + (i // 2) * HALF_STEP
        ws.add_image(img, f"{col}{r}")


def export_excel(M: pd.DataFrame, entity: str, statements: dict, audit,
                 ai_summary: str | None = None, ai_model: str | None = None,
                 view: str = "detailed") -> bytes:
    """Build the take-home workbook in memory — no filesystem writes, so this
    works unmodified on Streamlit Cloud.

    Sheet order mirrors the dashboard: Cover, (AI Summary), Executive Summary,
    the four Vernimmen stage sheets, then the Data Quality Log and Line Item
    Mapping. Every sheet keeps its full data table regardless of `view`; what
    changes is which chart images (PNG renders of the actual dashboard
    Plotly charts) get embedded:
      - "detailed": the full Vernimmen-stage breakdown charts, one set per
        stage sheet — matches the dashboard's Detailed tabs.
      - "simple": the four plain-language charts on the Executive Summary
        sheet only; stage sheets stay tables-only, matching the dashboard's
        Simple view (which has no per-stage chart breakdown either).
    The AI Summary sheet is included only when `ai_summary` is provided.
    """
    view = "simple" if str(view).lower().startswith("simple") else "detailed"
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    summ = executive_summary(M, entity)
    section_cols = {}

    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        summ.to_excel(xw, sheet_name="Executive Summary", index=False)
        for tab, cols in STAGE_MAP.items():
            cols = [c for c in cols if c in M.columns]
            section_cols[tab] = cols
            M[cols].T.to_excel(xw, sheet_name=tab[:31])
        pd.DataFrame(audit.entries or [{"Severity": "INFO", "Area": "-", "Finding": "No issues raised."}]) \
            .to_excel(xw, sheet_name="Data Quality Log", index=False)
        mapping = pd.DataFrame(
            [{"Statement": k, "Concept": c, "Matched line in your file": v}
             for k, st in statements.items() for c, v in st.resolved.items()])
        mapping.to_excel(xw, sheet_name="Line Item Mapping", index=False)

    buf.seek(0)

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(buf.getvalue()))
    navy = PatternFill("solid", fgColor="00182B")
    pale = PatternFill("solid", fgColor="F0F6FD")
    thin = Border(bottom=Side(style="thin", color="C1DDF9"))

    data_sheets = list(wb.sheetnames)  # sheets that carry a header row + tables

    # Header + number formatting on the tabular sheets (skip cover/AI, added later).
    for name in data_sheets:
        ws = wb[name]
        ws.freeze_panes = "B2"
        for cell in ws[1]:
            cell.fill = navy
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 22
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(size=10)
                cell.border = thin
                if isinstance(cell.value, float):
                    hdr = str(ws.cell(row=cell.row, column=1).value or "")
                    if "%" in hdr or "pts" in hdr:
                        cell.number_format = '#,##0.0"%"'
                    elif "(x)" in hdr or "Ratio" in hdr:
                        cell.number_format = '#,##0.00"x"'
                    elif "days" in hdr.lower():
                        cell.number_format = "#,##0.0"
                    else:
                        cell.number_format = '$#,##0;[Red]($#,##0)'
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(width + 2, 12), 52)

    ws = wb["Executive Summary"]
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 46
    for row in ws.iter_rows(min_row=2, max_col=1):
        c = row[0]
        if isinstance(c.value, str) and c.value.startswith("--"):
            for cc in ws[c.row]:
                cc.fill = pale
                cc.font = Font(bold=True, color="00182B", size=11)

    # Chart images: same Plotly figures as the dashboard, rendered to PNG.
    # Lazy import avoids a circular import (lib.charts imports money/pct from
    # this module).
    from . import charts as _charts

    if view == "detailed":
        for tab, fn_names in DETAIL_CHART_FUNCS.items():
            if tab not in wb.sheetnames:
                continue
            cols = section_cols.get(tab, [])
            figs = [getattr(_charts, fn)(M) for fn in fn_names]
            anchor_row = len(cols) + 4
            _embed_chart_images(wb[tab], anchor_row, figs)
    else:
        figs = [getattr(_charts, fn)(M) for fn in SIMPLE_CHART_FUNCS]
        anchor_row = len(summ) + 4
        _embed_chart_images(wb["Executive Summary"], anchor_row, figs, full_width_first=True)

    # Cover sheet first; AI summary sheet second (when present).
    has_ai = bool(ai_summary and ai_summary.strip())
    cover = wb.create_sheet("Cover", 0)
    _write_cover(cover, entity, M, has_ai, view)
    if has_ai:
        ai_ws = wb.create_sheet("AI Summary", 1)
        _write_ai_sheet(ai_ws, ai_summary, ai_model or "", entity)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
